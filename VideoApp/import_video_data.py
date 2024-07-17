import os
import openpyxl
from django.core.wsgi import get_wsgi_application

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'DDFWSql2.settings')
application = get_wsgi_application()

from VideoApp.models import *

def import_video_data_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        video_place = VideoPlace(
            information_address=row[0] if row[0] else '',
            copyright_owner=row[1] if row[1] else '',
            special_name=row[2] if row[2] else '',
            place_location_level1=PlaceLocation.objects.filter(name=row[3]).first() if row[3] else None,
            place_location_level2=PlaceLocation.objects.filter(name=row[4]).first() if row[4] else None,
            place_location_level3=PlaceLocation.objects.filter(name=row[5]).first() if row[5] else None,
            place_natural_level1=PlaceNatural.objects.filter(name=row[6]).first() if row[6] else None,
            place_natural_level2=PlaceNatural.objects.filter(name=row[7]).first() if row[7] else None,
            place_natural_level3=PlaceNatural.objects.filter(name=row[8]).first() if row[8] else None,
            place_humanistic_level1=PlaceHumanistic.objects.filter(name=row[9]).first() if row[9] else None,
            place_humanistic_level2=PlaceHumanistic.objects.filter(name=row[10]).first() if row[10] else None,
            place_humanistic_level3=PlaceHumanistic.objects.filter(name=row[11]).first() if row[11] else None,
            dir_name=row[12] if row[12] else '',
            season=row[13] if row[13] else '',
            video_class=row[14] if row[14] else '',
            picture_frame=row[15] if row[15] else '',
            keywords=row[16] if row[16] else '',
        )
        if video_place.information_address != '':
            print(video_place.information_address)
        else:
            print('empty')
        #video_place.save()

def import_video_humanistic(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.worksheets[1]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:  # 只有当information_address非空的时候才写入
            video_humanistic = VideoHumanistic(
                information_address=row[0],
                copyright_owner=row[1] if row[1] else '',
                special_name=row[2] if row[2] else '',
                humanistic_category_level1=HumanisticCategory.objects.filter(name=row[3]).first() if row[3] else None,
                humanistic_category_level2=HumanisticCategory.objects.filter(name=row[4]).first() if row[4] else None,
                humanistic_heritage_level1=HumanisticHeritage.objects.filter(name=row[5]).first() if row[5] else None,
                humanistic_heritage_level2=HumanisticHeritage.objects.filter(name=row[6]).first() if row[6] else None,
                humanistic_topic_level1=HumanisticTopic.objects.filter(name=row[7]).first() if row[7] else None,
                humanistic_topic_level2=HumanisticTopic.objects.filter(name=row[8]).first() if row[8] else None,
                dir_name=row[9] if row[9] else '',
                season=row[10] if row[10] else '',
                video_class=row[11] if row[11] else '',
                picture_frame=row[12] if row[12] else '',
                keywords=row[13] if row[13] else '',
            )
            print(video_humanistic.information_address)
            video_humanistic.save()


def import_video_cultural(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.worksheets[2]  # 获取第三个表

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:  # 只有当information_address非空的时候才写入
            video_cultural = VideoCultural(
                information_address=row[0],
                copyright_owner=row[1] if row[1] else '',
                special_name=row[2] if row[2] else '',
                cultural_agri_product_level1=CulturalAgriProduct.objects.filter(name=row[3]).first() if row[
                    3] else None,
                cultural_agri_product_level2=CulturalAgriProduct.objects.filter(name=row[4]).first() if row[
                    4] else None,
                cultural_delicacies_level1=CulturalDelicacies.objects.filter(name=row[5]).first() if row[5] else None,
                cultural_delicacies_level2=CulturalDelicacies.objects.filter(name=row[6]).first() if row[6] else None,
                cultural_processed_product_level1=CulturalProcessedProduct.objects.filter(name=row[7]).first() if row[
                    7] else None,
                cultural_processed_product_level2=CulturalProcessedProduct.objects.filter(name=row[8]).first() if row[
                    8] else None,
                dir_name=row[9] if row[9] else '',
                season=row[10] if row[10] else '',
                video_class=row[11] if row[11] else '',
                picture_frame=row[12] if row[12] else '',
                keywords=row[13] if row[13] else '',
            )
            video_cultural.save()


if __name__ == "__main__":
    file_path = "E:/WangTR2/College2/C2-3/DDFW/需求0716/0716视频数据测试更新版.xlsx"  # Update with your actual file path
    #import_video_data_from_excel(file_path)
    #import_video_humanistic(file_path)
    import_video_cultural(file_path)
